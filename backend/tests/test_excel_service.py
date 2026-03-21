import pytest
import os
import csv
from openpyxl import Workbook
from app.services.excel_service import excel_service
from app.services import doc_retrieval

FIXTURES_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "fixtures/excel"))

def test_excel_profile_basic():
    path = "basic.xlsx"
    result = excel_service.profile(path, root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert len(result["sheets"]) == 1
    sheet = result["sheets"][0]
    assert sheet["name"] == "Sales"
    assert sheet["used_range"] == "A1:C3"
    assert 1 in sheet["header_rows"]

def test_excel_profile_irregular():
    path = "irregular.xlsx"
    result = excel_service.profile(path, root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert len(result["sheets"]) == 2
    q1 = next(s for s in result["sheets"] if s["name"] == "Q1")
    assert "A1:C1" in q1["merged_cells"]
    assert "B" in q1["hidden_cols"]
    assert 4 in q1["hidden_rows"]

def test_excel_logic_extract():
    path = "logic.xlsx"
    result = excel_service.logic_extract(path, root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert len(result["formulas"]) >= 2
    assert any(f["cell"] == "A3" and "=SUM(A1:A2)" in f["formula"] for f in result["formulas"])
    assert any(f["cell"] == "B1" and "=A1*2" in f["formula"] for f in result["formulas"])
    
    # Lineage check
    lineage = result["lineage"]
    assert "A3" in lineage["nodes"]
    assert "A1" in lineage["nodes"]
    assert ["A1", "A3"] in lineage["edges"]
    
    # Named range check
    assert any(nr["name"] == "MyConst" for nr in result["named_ranges"])
    
    # Validation check
    assert len(result["validations"]) >= 1
    assert any(v["type"] == "whole" for v in result["validations"])

def test_excel_script_scan_no_macro():
    path = "basic.xlsx"
    result = excel_service.script_scan(path, root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert result["has_macro"] is False
    assert result["risk_level"] == "low"

def test_excel_read_json():
    path = "basic.xlsx"
    result = excel_service.read(path, root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert result["tool"] == "excel_read"
    assert len(result["data"]) == 2
    assert result["data"][0]["Product"] == "Apple"

def test_excel_read_markdown():
    path = "basic.xlsx"
    result = excel_service.read(path, mode="markdown", root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert isinstance(result["data"], str)
    assert "| Date" in result["data"]

def test_excel_query_basic():
    path = "basic.xlsx"
    query = "SELECT * FROM excel_data WHERE Amount > 120"
    result = excel_service.query(path, query, root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert len(result["data"]) == 1
    assert result["data"][0]["Product"] == "Banana"

def test_excel_query_invalid():
    path = "basic.xlsx"
    query = "DELETE FROM excel_data"
    result = excel_service.query(path, query, root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is False
    assert result["error_code"] == "INVALID_QUERY"

def test_excel_security_denied_path():
    with pytest.raises(doc_retrieval.DocAccessError):
        excel_service.profile("/etc/passwd")

def test_excel_security_allowed_roots():
    # Test path resolution with custom roots
    path = "basic.xlsx"
    result = excel_service.profile(path, root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True

def test_excel_profile_csv():
    csv_path = os.path.join(FIXTURES_DIR, "test.csv")
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Name", "Score"])
        writer.writerow([1, "Alice", 90])
        writer.writerow([2, "Bob", 85])
    
    result = excel_service.profile("test.csv", root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert result["sheets"][0]["name"] == "default"
    assert result["sheets"][0]["used_range"] == "A1:C3"

def test_excel_read_csv():
    result = excel_service.read("test.csv", root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert len(result["data"]) == 2
    assert result["data"][0]["Name"] == "Alice"

def test_excel_query_csv():
    query = "SELECT * FROM excel_data WHERE Score > 88"
    result = excel_service.query("test.csv", query, root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert len(result["data"]) == 1
    assert result["data"][0]["Name"] == "Alice"

def test_excel_profile_empty_sheet():
    path = os.path.join(FIXTURES_DIR, "empty.xlsx")
    wb = Workbook()
    wb.save(path)
    
    result = excel_service.profile("empty.xlsx", root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert result["sheets"][0]["header_rows"] == []

def test_excel_read_truncation():
    path = os.path.join(FIXTURES_DIR, "large.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["Val"])
    for i in range(600): # read_limit is 500
        ws.append([i])
    wb.save(path)
    
    result = excel_service.read("large.xlsx", root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert result["is_truncated"] is True
    assert len(result["data"]) == 500

def test_excel_multibyte_support():
    path = os.path.join(FIXTURES_DIR, "chinese.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(["日期", "产品", "金额"])
    ws.append(["2024-01-01", "苹果", 100])
    wb.save(path)
    
    result = excel_service.read("chinese.xlsx", root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert result["data"][0]["产品"] == "苹果"

def test_excel_formula_error_handling():
    path = os.path.join(FIXTURES_DIR, "error_formula.xlsx")
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 10
    ws["A2"] = 0
    ws["A3"] = "=A1/A2" # #DIV/0!
    wb.save(path)
    
    result = excel_service.read("error_formula.xlsx", root_dir=FIXTURES_DIR, allow_roots=[FIXTURES_DIR])
    assert result["ok"] is True
    assert "data" in result
