import os
import logging
import zipfile
import re
import csv
import duckdb
import time
import pandas as pd
from typing import List, Dict, Any, Optional, Tuple, Set
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from . import doc_retrieval

logger = logging.getLogger(__name__)

EXCEL_EXTENSIONS = [".xlsx", ".xlsm", ".xltx", ".xltm"]

class ExcelService:
    def __init__(self):
        self.max_rows = 100000  # Increased for performance tuning task later
        self.max_cols = 1000
        self.timeout_s = 30
        self.read_limit = 500  # Default max rows for read

    def _log_audit(self, tool: str, path: str, elapsed: float, status: str = "success", **kwargs):
        """Standardized audit logging for Excel tools."""
        audit_info = {
            "tool": tool,
            "file": os.path.basename(path),
            "elapsed_ms": round(elapsed * 1000, 2),
            "status": status,
            **kwargs
        }
        logger.info(f"[ExcelAudit] {audit_info}")

    def _resolve_path(self, path: str, root_dir: Optional[str] = None, allow_roots: Optional[List[str]] = None, deny_roots: Optional[List[str]] = None) -> str:
        """Resolve and validate the path using doc_retrieval logic."""
        return doc_retrieval.resolve_docs_path(
            path,
            docs_root=doc_retrieval.resolve_docs_root_for_read(
                path,
                requested_root=root_dir,
                allow_roots=allow_roots,
                deny_roots=deny_roots,
                allowed_extensions=EXCEL_EXTENSIONS + [".csv"],
            ),
            require_md=False,
            allowed_extensions=EXCEL_EXTENSIONS + [".csv"]
        )

    def _profile_csv(self, abs_path: str) -> Dict[str, Any]:
        """Basic CSV profile."""
        with open(abs_path, 'r', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            rows = list(reader)
            row_count = len(rows)
            col_count = len(rows[0]) if rows else 0
            
        return {
            "ok": True,
            "tool": "excel_profile",
            "file": abs_path,
            "sheets": [{
                "name": "default",
                "used_range": f"A1:{get_column_letter(col_count)}{row_count}",
                "header_rows": [1] if row_count > 0 else [],
                "data_blocks": [{"start_row": 2, "end_row": row_count, "start_col": 1, "end_col": col_count}] if row_count > 1 else [],
                "merged_cells": [],
                "hidden_rows": [],
                "hidden_cols": []
            }]
        }

    def profile(self, path: str, root_dir: Optional[str] = None, allow_roots: Optional[List[str]] = None, deny_roots: Optional[List[str]] = None, request_id: Optional[str] = None) -> Dict[str, Any]:
        """Identify Excel structure: header rows, data blocks, merged cells, etc."""
        start_time = time.time()
        try:
            abs_path = self._resolve_path(path, root_dir, allow_roots, deny_roots)
            
            if abs_path.lower().endswith(".csv"):
                result = self._profile_csv(abs_path)
                self._log_audit("excel_profile", path, time.time() - start_time, request_id=request_id)
                return result

            # For profile, we use read_only=False to get full metadata like merged_cells
            # but we still use data_only=True for value calculation if needed.
            wb = load_workbook(abs_path, read_only=False, data_only=True)
            sheets_info = []
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Basic info
                used_range = ws.calculate_dimension()
                
                # Hidden rows/cols - now we can access these with read_only=False
                hidden_rows = [r for r, dim in ws.row_dimensions.items() if getattr(dim, 'hidden', False)]
                hidden_cols = []
                for c, dim in ws.column_dimensions.items():
                    if getattr(dim, 'hidden', False):
                        if isinstance(c, int):
                            hidden_cols.append(get_column_letter(c))
                        else:
                            hidden_cols.append(str(c))
                
                sheets_info.append({
                    "name": sheet_name,
                    "used_range": used_range,
                    "hidden_rows": hidden_rows,
                    "hidden_cols": hidden_cols,
                    "merged_cells": [str(m) for m in ws.merged_cells.ranges] if hasattr(ws, 'merged_cells') else [],
                    "header_rows": [],
                    "data_blocks": []
                })
                
                # Simple header inference: first non-empty row
                headers = []
                for row_idx, row in enumerate(ws.iter_rows(max_row=10), 1):
                    if any(cell.value is not None for cell in row):
                        headers.append(row_idx)
                        break
                sheets_info[-1]["header_rows"] = headers
                
                # Simple data block inference: everything after header
                if headers:
                    min_row = headers[-1] + 1
                    max_row = ws.max_row if ws.max_row else 0
                    min_col = 1
                    max_col = ws.max_column if ws.max_column else 0
                    if max_row >= min_row:
                        sheets_info[-1]["data_blocks"] = [{
                            "start_row": min_row,
                            "end_row": max_row,
                            "start_col": min_col,
                            "end_col": max_col
                        }]
            
            wb.close()
            
            elapsed = time.time() - start_time
            self._log_audit("excel_profile", path, elapsed, request_id=request_id, sheets_count=len(sheets_info))
            
            return {
                "ok": True,
                "tool": "excel_profile",
                "file": path,
                "sheets": sheets_info
            }
        except Exception as e:
            elapsed = time.time() - start_time
            self._log_audit("excel_profile", path, elapsed, status="failed", error=str(e), request_id=request_id)
            raise e

    def logic_extract(self, path: str, sheet_name: Optional[str] = None, root_dir: Optional[str] = None, allow_roots: Optional[List[str]] = None, deny_roots: Optional[List[str]] = None, request_id: Optional[str] = None) -> Dict[str, Any]:
        """Extract formulas, lineage, named ranges, validations, connections."""
        start_time = time.time()
        try:
            abs_path = self._resolve_path(path, root_dir, allow_roots, deny_roots)
            
            # Load with formulas (data_only=False) - logic_extract requires this
            wb = load_workbook(abs_path, data_only=False)
            
            target_sheets = [sheet_name] if sheet_name else wb.sheetnames
            
            formulas = []
            lineage_nodes = set()
            lineage_edges = []
            
            for s_name in target_sheets:
                if s_name not in wb.sheetnames:
                    continue
                ws = wb[s_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            cell_ref = f"{cell.coordinate}"
                            formula = cell.value
                            formulas.append({"cell": cell_ref, "formula": formula})
                            
                            # Simple lineage extraction (regex based)
                            refs = re.findall(r'([A-Z]+\d+|[A-Z]+\$\d+|\$[A-Z]+\d+|\$[A-Z]+\$\d+)', formula)
                            for ref in refs:
                                ref_clean = ref.replace('$', '')
                                lineage_nodes.add(cell_ref)
                                lineage_nodes.add(ref_clean)
                                lineage_edges.append([ref_clean, cell_ref])

            named_ranges = []
            for name, defn in wb.defined_names.items():
                named_ranges.append({
                    "name": name,
                    "refers_to": str(defn.attr_text) if hasattr(defn, 'attr_text') else str(defn)
                })

            validations = []
            for s_name in target_sheets:
                if s_name not in wb.sheetnames:
                    continue
                ws = wb[s_name]
                for dv in ws.data_validations.dataValidation:
                    validations.append({
                        "range": str(dv.sqref),
                        "type": dv.type,
                        "formula1": dv.formula1
                    })

            connections = []
            if hasattr(wb, '_external_links'):
                for link in wb._external_links:
                    connections.append(str(link))

            wb.close()
            
            elapsed = time.time() - start_time
            self._log_audit("excel_logic_extract", path, elapsed, request_id=request_id, formulas_count=len(formulas))
            
            return {
                "ok": True,
                "tool": "excel_logic_extract",
                "sheet": sheet_name or "all",
                "formulas": formulas[:100],
                "lineage": {
                    "nodes": list(lineage_nodes),
                    "edges": lineage_edges
                },
                "named_ranges": named_ranges,
                "validations": validations,
                "connections": connections
            }
        except Exception as e:
            elapsed = time.time() - start_time
            self._log_audit("excel_logic_extract", path, elapsed, status="failed", error=str(e), request_id=request_id)
            raise e

    def script_scan(self, path: str, root_dir: Optional[str] = None, allow_roots: Optional[List[str]] = None, deny_roots: Optional[List[str]] = None, request_id: Optional[str] = None) -> Dict[str, Any]:
        """Static scan for VBA/macros."""
        start_time = time.time()
        try:
            abs_path = self._resolve_path(path, root_dir, allow_roots, deny_roots)
            
            has_macro = False
            risk_level = "low"
            hits = []
            
            if zipfile.is_zipfile(abs_path):
                with zipfile.ZipFile(abs_path, 'r') as z:
                    file_list = z.namelist()
                    if 'xl/vbaProject.bin' in file_list:
                        has_macro = True
                        risk_level = "medium"
                        try:
                            with z.open('xl/vbaProject.bin') as f:
                                content = f.read()
                                keywords = [
                                    b"Shell", b"CreateObject", b"WScript", b"PowerShell", 
                                    b"ADODB", b"FileSystemObject", b"Environ", b"Declare"
                                ]
                                for kw in keywords:
                                    count = content.count(kw)
                                    if count > 0:
                                        hits.append({"keyword": kw.decode('ascii'), "count": count})
                                        risk_level = "high"
                        except Exception as e:
                            logger.warning(f"Failed to scan vbaProject.bin: {e}")
            
            summary = "No macros detected."
            if has_macro:
                if risk_level == "high":
                    summary = "Detected potentially dangerous macro patterns. Execution is blocked by policy."
                else:
                    summary = "Macros detected, but no highly suspicious patterns found. Execution is blocked by policy."

            elapsed = time.time() - start_time
            self._log_audit("excel_script_scan", path, elapsed, request_id=request_id, has_macro=has_macro, risk_level=risk_level)
            
            return {
                "ok": True,
                "tool": "excel_script_scan",
                "has_macro": has_macro,
                "risk_level": risk_level,
                "hits": hits,
                "summary": summary,
                "action": "static-analysis-only"
            }
        except Exception as e:
            elapsed = time.time() - start_time
            self._log_audit("excel_script_scan", path, elapsed, status="failed", error=str(e), request_id=request_id)
            raise e

    def read(self, path: str, sheet_name: Optional[str] = None, range: Optional[str] = None, mode: str = "json", root_dir: Optional[str] = None, allow_roots: Optional[List[str]] = None, deny_roots: Optional[List[str]] = None, request_id: Optional[str] = None) -> Dict[str, Any]:
        """Read data from Excel with truncation and mode support."""
        start_time = time.time()
        try:
            abs_path = self._resolve_path(path, root_dir, allow_roots, deny_roots)
            
            if abs_path.lower().endswith(".csv"):
                df = pd.read_csv(abs_path)
                sheet_name = "default"
            else:
                wb = load_workbook(abs_path, read_only=True, data_only=True)
                if not sheet_name:
                    sheet_name = wb.sheetnames[0]
                ws = wb[sheet_name]
                
                if range:
                    min_col, min_row, max_col, max_row = range_boundaries(range)
                    data = ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
                    df = pd.DataFrame(data)
                else:
                    data = ws.values
                    cols = next(data)
                    df = pd.DataFrame(data, columns=cols)
                wb.close()

            # Truncation
            total_rows = len(df)
            df = df.head(self.read_limit)
            is_truncated = total_rows > self.read_limit
            
            if mode == "markdown":
                try:
                    content = df.to_markdown(index=False)
                except ImportError:
                    columns = [str(col) for col in df.columns]
                    header = "| " + " | ".join(columns) + " |"
                    separator = "| " + " | ".join(["---"] * len(columns)) + " |"
                    rows = []
                    for _, row in df.iterrows():
                        values = ["" if pd.isna(v) else str(v) for v in row.tolist()]
                        rows.append("| " + " | ".join(values) + " |")
                    content = "\n".join([header, separator, *rows])
            else:
                content = df.to_dict(orient="records")

            elapsed = time.time() - start_time
            self._log_audit("excel_read", path, elapsed, request_id=request_id, is_truncated=is_truncated, total_rows=total_rows)
            
            return {
                "ok": True,
                "tool": "excel_read",
                "file": path,
                "sheet": sheet_name,
                "range": range,
                "is_truncated": is_truncated,
                "total_rows": total_rows,
                "data": content
            }
        except Exception as e:
            elapsed = time.time() - start_time
            self._log_audit("excel_read", path, elapsed, status="failed", error=str(e), request_id=request_id)
            raise e

    def query(self, path: str, query: str, sheet_name: Optional[str] = None, root_dir: Optional[str] = None, allow_roots: Optional[List[str]] = None, deny_roots: Optional[List[str]] = None, request_id: Optional[str] = None) -> Dict[str, Any]:
        """Execute SQL query on Excel data using DuckDB."""
        start_time = time.time()
        try:
            abs_path = self._resolve_path(path, root_dir, allow_roots, deny_roots)
            
            if abs_path.lower().endswith(".csv"):
                df = pd.read_csv(abs_path)
            else:
                # OPTIMIZATION: use read_only=True for memory efficiency
                wb = load_workbook(abs_path, read_only=True, data_only=True)
                if not sheet_name:
                    sheet_name = wb.sheetnames[0]
                ws = wb[sheet_name]
                data = ws.values
                cols = next(data)
                df = pd.DataFrame(data, columns=cols)
                wb.close()

            # DuckDB in-memory query
            con = duckdb.connect(database=':memory:')
            # Register the dataframe as a table
            con.register('excel_data', df)
            
            # Security: Only allow SELECT
            if not re.match(r'^\s*SELECT', query, re.IGNORECASE):
                self._log_audit("excel_query", path, time.time() - start_time, status="blocked", reason="non-select-query", request_id=request_id)
                return {
                    "ok": False,
                    "error_code": "INVALID_QUERY",
                    "message": "Only SELECT queries are allowed.",
                    "hint": "Ensure your query starts with SELECT."
                }

            res_df = con.execute(query).df()
            total_results = len(res_df)
            res_df = res_df.head(self.read_limit)
            is_truncated = total_results > self.read_limit
            
            elapsed = time.time() - start_time
            self._log_audit("excel_query", path, elapsed, request_id=request_id, query=query[:100], total_results=total_results, is_truncated=is_truncated)
            
            return {
                "ok": True,
                "tool": "excel_query",
                "file": path,
                "query": query,
                "is_truncated": is_truncated,
                "total_results": total_results,
                "data": res_df.to_dict(orient="records")
            }
        except Exception as e:
            elapsed = time.time() - start_time
            self._log_audit("excel_query", path, elapsed, status="failed", error=str(e), request_id=request_id)
            return {
                "ok": False,
                "error_code": "QUERY_EXECUTION_FAILED",
                "message": str(e),
                "hint": "Check your SQL syntax and column names."
            }
        finally:
            if 'con' in locals():
                con.close()

excel_service = ExcelService()
